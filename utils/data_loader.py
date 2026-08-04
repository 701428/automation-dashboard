"""
Data loading layer — reads directly from Automation tracker.xlsx.

Flow:
  1. Auto-detect tracker in project root or data/ dir; if newer than internal store → re-parse
  2. All load_* functions read from MAIN_FILE (openpyxl, stable format)
  3. save_* functions write back to MAIN_FILE (user overrides persist between tracker uploads)
  4. process_uploaded_file() detects tracker sheets → saves as TRACKER_FILE → triggers re-parse
  5. project_config.json stores runtime-only metadata (color, priority, team_size) per project
"""

import io
import json
import re
from pathlib import Path
from typing import Optional

import pandas as pd
import numpy as np

DATA_DIR     = Path(__file__).parent.parent / "data"
MAIN_FILE    = DATA_DIR / "automation_data.xlsx"
TRACKER_FILE = DATA_DIR / "Automation tracker.xlsx"
CONFIG_FILE  = DATA_DIR / "project_config.json"
UPLOAD_DIR   = DATA_DIR / "uploads"
ROOT_TRACKER = Path(__file__).parent.parent / "Automation tracker.xlsx"

DATA_DIR.mkdir(exist_ok=True)
UPLOAD_DIR.mkdir(exist_ok=True)

_DEFAULT_PALETTE = [
    "#1645a4", "#02c9a8", "#37aafe", "#7c3aed", "#0891b2",
    "#db2777", "#ea580c", "#059669", "#f59e0b", "#64748b",
    "#dc2626", "#0d9488", "#7c3aed", "#0369a1",
]


# ── Slug / ID helpers ──────────────────────────────────────────────────────────

def _slugify(name: str) -> str:
    """'HES-Sangai' → 'hes_sangai', 'Meter Setu' → 'meter_setu'"""
    s = name.lower()
    s = re.sub(r'[\s\-/()+]+', '_', s)
    s = re.sub(r'[^a-z0-9_]', '', s)
    return re.sub(r'_+', '_', s).strip('_')


def _firmware_sub_id(name: str) -> Optional[str]:
    """Map firmware completion-plan row name to a stable project ID."""
    n = name.lower()
    if "overall" in n:
        return None  # skip the aggregate row
    if "1-phase" in n or "1 phase" in n:
        return "1p"
    if ("3-phase" in n or "3phase" in n) and "wc" in n:
        return "3p_wc"
    if ("3-phase" in n or "3phase" in n) and "ltct" in n:
        return "3p_ltct"
    return _slugify(name)


def _firmware_env_prefix(pid: str) -> str:
    return {"1p": "1-phase", "3p_wc": "3-phase wc", "3p_ltct": "3-phase ltct"}.get(pid, pid)


# ── Project config JSON ────────────────────────────────────────────────────────

def _load_project_config() -> dict:
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE) as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_project_config(cfg: dict) -> None:
    with open(CONFIG_FILE, 'w') as f:
        json.dump(cfg, f, indent=2)


def _project_cfg(pid: str, cfg: dict, idx: int = 0) -> dict:
    """Return config for pid, creating defaults if absent."""
    if pid not in cfg:
        cfg[pid] = {
            "color":     _DEFAULT_PALETTE[idx % len(_DEFAULT_PALETTE)],
            "priority":  "Medium",
            "team_size": 1,
            "daily_avg": 0,
        }
    return cfg[pid]


# ── Parsing helpers ────────────────────────────────────────────────────────────

def _parse_int(val) -> int:
    if val is None or (isinstance(val, float) and (val != val)):
        return 0
    try:
        return int(float(str(val).strip().split('(')[0].replace(',', '')))
    except Exception:
        m = re.search(r"\d+", str(val))
        return int(m.group()) if m else 0


def _parse_float(val) -> float:
    if val is None or (isinstance(val, float) and (val != val)):
        return 0.0
    try:
        return float(re.search(r"[\d.]+", str(val)).group())
    except Exception:
        return 0.0


def _parse_date(val) -> Optional[str]:
    if val is None:
        return None
    try:
        if isinstance(val, float) and val != val:
            return None
        ts = pd.to_datetime(val, errors="coerce")
        if pd.notna(ts):
            return ts.date().isoformat()
    except Exception:
        pass
    s = str(val).strip()
    return s if s.lower() not in ("nan", "nat", "none", "tbd", "", "n/a") else None


def _safe_str(val) -> str:
    if val is None or (isinstance(val, float) and val != val):
        return ""
    return str(val).strip()


def _find_row(df: pd.DataFrame, keyword: str) -> Optional[int]:
    for i, val in enumerate(df.iloc[:, 0]):
        if isinstance(val, str) and keyword.lower() in val.lower():
            return i
    return None


def _rows_between(df: pd.DataFrame, start: Optional[int], end: Optional[int]) -> list:
    """Return raw rows from start+2 to end (exclusive), skipping all-NaN rows."""
    if start is None:
        return []
    data_start = start + 2
    data_end   = end if end is not None else len(df)
    rows = []
    for i in range(data_start, data_end):
        row = df.iloc[i]
        if not row.isna().all():
            rows.append(row.tolist())
    return rows


def _count_from_name(name: str) -> int:
    m = re.search(r"\((\d+)\s*[Tt]est", name)
    if m:
        return int(m.group(1))
    parts = [p.strip() for p in name.split(",") if p.strip()]
    return len(parts) if len(parts) > 1 else 1


def _normalize_status(raw: str) -> str:
    r = raw.lower().strip()
    if not r or r in ("nan", ""):
        return "In Progress"
    if r in ("scheduled", "not started", "not started yet", "planning pending"):
        return "Not Started"
    if r in ("done", "completed"):
        return "Completed"
    if "on hold" in r:
        return "On Hold"
    if r == "delayed":
        return "Delayed"
    if "on track" in r:
        return "On Track"
    if "in progress" in r or "progress" in r:
        return "In Progress"
    return raw.strip()


def _has_tracker_format(df: pd.DataFrame) -> bool:
    """Return True if df looks like a tracker sheet (has our section headers in col 0)."""
    for val in df.iloc[:5, 0]:
        if isinstance(val, str) and "automation status summary" in val.lower():
            return True
    return False


# ── Firmware sheet parser ──────────────────────────────────────────────────────

def _parse_firmware_sheet(df, summary_row, na_rows, dp_rows, comp_rows,
                           proj_out, na_out, plan_out, comp_out, cfg, base_idx):
    """Parse Firmware sheet — sub-projects discovered dynamically from completion plan."""
    n_added = 0

    for r in comp_rows:
        name = _safe_str(r[0])
        if not name or name.lower() in ("project / environment", "nan"):
            continue

        pid = _firmware_sub_id(name)
        if pid is None:
            # Write Overall row to completion plan but skip as a project
            comp_out.append({
                "project_id":          "firmware_overall",
                "name":                name,
                "total_cases":         _parse_int(r[1]),
                "automatable":         _parse_int(r[1]),
                "duration_days":       _parse_int(r[2]),
                "daily_avg":           _parse_float(r[3]),
                "start_date":          _parse_date(r[4]) or _safe_str(r[4]) or "TBD",
                "expected_completion": _parse_date(r[5]) or _safe_str(r[5]) or "TBD",
                "status":              _safe_str(r[6]) if len(r) > 6 else "",
            })
            continue

        pcfg = _project_cfg(pid, cfg, base_idx + n_added)

        total       = _parse_int(r[1])
        raw_avg     = _parse_float(r[3])
        daily_avg   = raw_avg if 0 < raw_avg <= 100 else pcfg.get("daily_avg", 0)
        start_date  = _parse_date(r[4]) or _safe_str(r[4]) or "TBD"
        target_date = _parse_date(r[5]) or _safe_str(r[5]) or "TBD"
        status      = _normalize_status(_safe_str(r[6]) if len(r) > 6 else "")

        # Automated: count day-plan rows that are NOT marked "Planned"
        env_prefix = _firmware_env_prefix(pid)
        automated  = 0
        for dp_r in dp_rows:
            env_cell = _safe_str(dp_r[1]).lower()
            if not env_cell.startswith(env_prefix[:6]):
                continue
            remarks = _safe_str(dp_r[6]).lower() if len(dp_r) > 6 else ""
            if "planned" in remarks:
                continue
            actual = _parse_int(dp_r[3])
            if actual > 0:
                automated = max(automated, _parse_int(dp_r[4]))

        proj_out.append({
            "id":              pid,
            "name":            name,
            "total_cases":     total,
            "automatable":     total,
            "non_automatable": 0,
            "automated":       automated,
            "in_progress":     0,
            "team_size":       pcfg.get("team_size", 1),
            "start_date":      start_date,
            "target_date":     target_date,
            "daily_avg":       daily_avg,
            "status":          status,
            "priority":        pcfg.get("priority", "High"),
            "color":           pcfg.get("color", _DEFAULT_PALETTE[(base_idx + n_added) % len(_DEFAULT_PALETTE)]),
            "notes":           "",
        })

        # Day plan
        for dp_r in dp_rows:
            env_cell = _safe_str(dp_r[1]).lower()
            if not env_cell.startswith(env_prefix[:6]):
                continue
            actual   = _parse_int(dp_r[3])
            remarks  = _safe_str(dp_r[6]) if len(dp_r) > 6 else ""
            plan_out.append({
                "project_id":    pid,
                "date":          _parse_date(dp_r[0]) or _safe_str(dp_r[0]),
                "module":        _safe_str(dp_r[2]),
                "planned_cases": actual,
                "actual_cases":  actual,
                "cumulative":    _parse_int(dp_r[4]),
                "assigned_to":   _safe_str(dp_r[5]) if len(dp_r) > 5 else "",
                "remarks":       remarks,
                "status":        "Planned" if "planned" in remarks.lower() else ("Completed" if actual > 0 else "Planned"),
            })

        # Completion plan row
        comp_out.append({
            "project_id":          pid,
            "name":                name,
            "total_cases":         total,
            "automatable":         total,
            "duration_days":       _parse_int(r[2]),
            "daily_avg":           _parse_float(r[3]),
            "start_date":          start_date,
            "expected_completion": target_date,
            "status":              _safe_str(r[6]) if len(r) > 6 else "",
        })

        n_added += 1

    # Non-automatable (firmware non-auto rows belong to 1p if it exists, else firmware)
    fw_na_pid = "1p" if any(p["id"] == "1p" for p in proj_out) else "firmware"
    for r in na_rows:
        name_val = _safe_str(r[0])
        if not name_val or name_val.lower() in ("test case id", "nan"):
            continue
        na_out.append({
            "project_id": fw_na_pid,
            "module":     name_val,
            "count":      _count_from_name(name_val),
            "reason":     _safe_str(r[2]) if len(r) > 2 else "",
            "approach":   _safe_str(r[3]) if len(r) > 3 else "",
        })

    return n_added


# ── Single-sheet parser ────────────────────────────────────────────────────────

def _parse_single_sheet(sheet, df, summary_row, na_rows, dp_rows, comp_rows,
                         proj_out, na_out, plan_out, comp_out, cfg, idx):
    pid  = _slugify(sheet)
    pcfg = _project_cfg(pid, cfg, idx)

    # Project name from summary row col 0, fallback to prettified sheet name
    name = re.sub(r'[_\-]+', ' ', sheet).strip()
    total = automated = non_auto = 0

    if summary_row is not None and summary_row + 2 < len(df):
        sr       = df.iloc[summary_row + 2]
        raw_name = _safe_str(sr.iloc[0])
        if raw_name and raw_name.lower() not in ("nan", ""):
            name = raw_name
        total     = _parse_int(sr.iloc[1])
        automated = _parse_int(sr.iloc[2])
        non_auto  = _parse_int(sr.iloc[4])

    automatable = max(0, total - non_auto)

    # Dates, daily_avg, and status from first valid completion plan row
    start_date = target_date = "TBD"
    daily_avg  = pcfg.get("daily_avg", 0)
    status     = "In Progress"
    for r in comp_rows:
        row_name = _safe_str(r[0])
        if not row_name or row_name.lower() in ("project / environment", "nan"):
            continue
        raw_avg    = _parse_float(r[3])
        # cap at 100: values > 100 are likely dates mistaken as numbers (e.g. WFM has 2026-05-26)
        daily_avg  = raw_avg if 0 < raw_avg <= 100 else pcfg.get("daily_avg", 0)
        start_raw  = _parse_date(r[4]) or _safe_str(r[4]) or "TBD"
        target_raw = _parse_date(r[5]) or _safe_str(r[5]) or "TBD"
        # Reject strings that are clearly notes, not dates (> 20 chars)
        start_date  = start_raw  if len(start_raw)  <= 20 else "TBD"
        target_date = target_raw if len(target_raw) <= 20 else "TBD"
        raw_st     = _safe_str(r[6]) if len(r) > 6 else ""
        # fallback: some sheets put status in col 2 (e.g. WFM "On Hold")
        if not raw_st or raw_st.lower() in ("nan", ""):
            raw_st = _safe_str(r[2]) if len(r) > 2 else ""
        status = _normalize_status(raw_st)
        break

    proj_out.append({
        "id":              pid,
        "name":            name,
        "total_cases":     total,
        "automatable":     automatable,
        "non_automatable": non_auto,
        "automated":       automated,
        "in_progress":     0,
        "team_size":       pcfg.get("team_size", 1),
        "start_date":      start_date,
        "target_date":     target_date,
        "daily_avg":       daily_avg,
        "status":          status,
        "priority":        pcfg.get("priority", "Medium"),
        "color":           pcfg.get("color", _DEFAULT_PALETTE[idx % len(_DEFAULT_PALETTE)]),
        "notes":           "",
    })

    # Non-automatable
    for r in na_rows:
        name_val = _safe_str(r[0])
        if not name_val or name_val.lower() in ("test case id", "nan"):
            continue
        na_out.append({
            "project_id": pid,
            "module":     name_val,
            "count":      _count_from_name(name_val),
            "reason":     _safe_str(r[2]) if len(r) > 2 else "",
            "approach":   _safe_str(r[3]) if len(r) > 3 else "",
        })

    # Day plan
    for r in dp_rows:
        if all(_safe_str(v) == "" for v in r[:5]):
            continue
        actual  = _parse_int(r[3])
        remarks = _safe_str(r[6]) if len(r) > 6 else ""
        plan_out.append({
            "project_id":    pid,
            "date":          _parse_date(r[0]) or _safe_str(r[0]),
            "module":        _safe_str(r[2]),
            "planned_cases": actual if actual > 0 else 0,
            "actual_cases":  actual,
            "cumulative":    _parse_int(r[4]),
            "assigned_to":   _safe_str(r[5]) if len(r) > 5 else "",
            "remarks":       remarks,
            "status":        "Planned" if "planned" in remarks.lower() else ("Completed" if actual > 0 else "Planned"),
        })

    # Completion plan
    for r in comp_rows:
        row_name = _safe_str(r[0])
        if not row_name or row_name.lower() in ("project / environment", "nan"):
            continue
        comp_out.append({
            "project_id":          pid,
            "name":                row_name,
            "total_cases":         _parse_int(r[1]),
            "automatable":         max(0, _parse_int(r[1]) - non_auto),
            "duration_days":       _parse_int(r[2]),
            "daily_avg":           _parse_float(r[3]),
            "start_date":          _parse_date(r[4]) or _safe_str(r[4]),
            "expected_completion": _parse_date(r[5]) or _safe_str(r[5]),
            "status":              _safe_str(r[6]) if len(r) > 6 else "",
        })


# ── Core tracker parser ────────────────────────────────────────────────────────

def _parse_and_save_tracker(tracker_path: Path) -> None:
    """Parse all tracker sheets dynamically and write to MAIN_FILE."""
    xl  = pd.ExcelFile(tracker_path)
    cfg = _load_project_config()

    all_projects  = []
    all_non_auto  = []
    all_day_plan  = []
    all_completion= []
    proj_idx      = 0

    for sheet in xl.sheet_names:
        df = xl.parse(sheet, header=None)
        if not _has_tracker_format(df):
            continue

        summary_row    = _find_row(df, "Automation Status Summary")
        nonaut_row     = _find_row(df, "Non-Automatable Test Cases")
        dayplan_row    = _find_row(df, "Day-by-Day Automation")
        completion_row = _find_row(df, "Project Completion Plan")

        na_rows   = _rows_between(df, nonaut_row,     dayplan_row)
        dp_rows   = _rows_between(df, dayplan_row,    completion_row)
        comp_rows = _rows_between(df, completion_row, None)

        if sheet == "Firmware":
            n = _parse_firmware_sheet(
                df, summary_row, na_rows, dp_rows, comp_rows,
                all_projects, all_non_auto, all_day_plan, all_completion,
                cfg, proj_idx,
            )
            proj_idx += n
        else:
            _parse_single_sheet(
                sheet, df, summary_row, na_rows, dp_rows, comp_rows,
                all_projects, all_non_auto, all_day_plan, all_completion,
                cfg, proj_idx,
            )
            proj_idx += 1

    _save_project_config(cfg)

    with pd.ExcelWriter(MAIN_FILE, engine="openpyxl") as writer:
        pd.DataFrame(all_projects).to_excel(   writer, sheet_name="Projects",        index=False)
        pd.DataFrame(all_non_auto).to_excel(   writer, sheet_name="Non_Automatable", index=False)
        pd.DataFrame(all_day_plan).to_excel(   writer, sheet_name="Day_Plan",        index=False)
        pd.DataFrame(all_completion).to_excel( writer, sheet_name="Completion_Plan", index=False)


# ── Fallback seed (used only when no tracker file exists at all) ───────────────

def _create_sample_excel() -> None:
    rows_proj = [
        {"id":"1p",            "name":"1-Phase Meter Firmware",       "total_cases":781,  "automatable":781, "non_automatable":0,  "automated":0,"in_progress":0,"team_size":1,"start_date":"2026-09-16","target_date":"2027-02-28","daily_avg":5, "status":"Not Started",    "priority":"High",  "color":"#1645a4","notes":""},
        {"id":"3p_wc",         "name":"3-Phase WC",                   "total_cases":454,  "automatable":454, "non_automatable":0,  "automated":0,"in_progress":0,"team_size":1,"start_date":"TBD",       "target_date":"TBD",       "daily_avg":0, "status":"Not Started",    "priority":"Medium","color":"#02c9a8","notes":""},
        {"id":"3p_ltct",       "name":"3-Phase LTCT",                 "total_cases":455,  "automatable":455, "non_automatable":0,  "automated":0,"in_progress":0,"team_size":1,"start_date":"TBD",       "target_date":"TBD",       "daily_avg":0, "status":"Not Started",    "priority":"Medium","color":"#37aafe","notes":""},
        {"id":"hes_sangai",    "name":"Sangai",                       "total_cases":164,  "automatable":158, "non_automatable":6,  "automated":81,"in_progress":0,"team_size":1,"start_date":"2026-06-19","target_date":"2026-08-18","daily_avg":3, "status":"Delayed",       "priority":"High",  "color":"#7c3aed","notes":""},
        {"id":"hes_goamti",    "name":"HES-Gomati",                   "total_cases":164,  "automatable":158, "non_automatable":6,  "automated":120,"in_progress":0,"team_size":1,"start_date":"2026-06-19","target_date":"2026-08-05","daily_avg":3,"status":"Delayed",       "priority":"High",  "color":"#0891b2","notes":""},
        {"id":"display_automation","name":"Smart Meter / Jig",        "total_cases":14,   "automatable":14,  "non_automatable":0,  "automated":14,"in_progress":0,"team_size":1,"start_date":"2026-06-22","target_date":"2026-08-03","daily_avg":2, "status":"Completed",     "priority":"Medium","color":"#db2777","notes":""},
        {"id":"vee",           "name":"VEE (Gomati / Sangai)",        "total_cases":603,  "automatable":571, "non_automatable":32, "automated":460,"in_progress":0,"team_size":1,"start_date":"2026-07-10","target_date":"2026-09-11","daily_avg":0,"status":"Delayed",       "priority":"High",  "color":"#ea580c","notes":""},
        {"id":"consumer_app",  "name":"Consumer App (Gomati Android)","total_cases":81,   "automatable":81,  "non_automatable":0,  "automated":17,"in_progress":0,"team_size":1,"start_date":"2026-07-03","target_date":"2026-09-30","daily_avg":1, "status":"In Progress",   "priority":"Medium","color":"#059669","notes":""},
        {"id":"comms",         "name":"Comms (4G / RF / IMG / DCU)", "total_cases":86,   "automatable":60,  "non_automatable":26, "automated":18,"in_progress":0,"team_size":1,"start_date":"2026-07-07","target_date":"2026-09-15","daily_avg":3, "status":"Delayed",       "priority":"High",  "color":"#f59e0b","notes":""},
        {"id":"wfm",           "name":"WFM Portal Stage",             "total_cases":278,  "automatable":123, "non_automatable":155,"automated":117,"in_progress":0,"team_size":1,"start_date":"TBD",       "target_date":"TBD",       "daily_avg":0,"status":"On Hold",       "priority":"Medium","color":"#64748b","notes":""},
        {"id":"meter_setu",    "name":"Meter Setu (Gomati)",          "total_cases":90,   "automatable":89,  "non_automatable":1,  "automated":0,"in_progress":0,"team_size":1,"start_date":"2026-08-04","target_date":"2026-09-30","daily_avg":2, "status":"Not Started",   "priority":"Medium","color":"#dc2626","notes":""},
    ]
    with pd.ExcelWriter(MAIN_FILE, engine="openpyxl") as writer:
        pd.DataFrame(rows_proj).to_excel(writer, sheet_name="Projects", index=False)
        pd.DataFrame(columns=["project_id","module","count","reason","approach"]
            ).to_excel(writer, sheet_name="Non_Automatable", index=False)
        pd.DataFrame(columns=["project_id","date","module","planned_cases","actual_cases",
                               "cumulative","assigned_to","remarks","status"]
            ).to_excel(writer, sheet_name="Day_Plan", index=False)
        pd.DataFrame(columns=["project_id","name","total_cases","automatable","duration_days",
                               "daily_avg","start_date","expected_completion","status"]
            ).to_excel(writer, sheet_name="Completion_Plan", index=False)


# ── Public API ─────────────────────────────────────────────────────────────────

def ensure_data_file() -> None:
    """
    Sync strategy:
      1. If root Automation tracker.xlsx exists and is newer than internal tracker → copy it
      2. If TRACKER_FILE exists and is newer than MAIN_FILE → re-parse tracker
      3. Else if MAIN_FILE missing → seed from defaults
    """
    # Auto-import tracker from project root if it's newer than what's in data/
    if ROOT_TRACKER.exists() and ROOT_TRACKER.stat().st_size > 1024:
        root_mtime    = ROOT_TRACKER.stat().st_mtime
        tracker_mtime = TRACKER_FILE.stat().st_mtime if TRACKER_FILE.exists() else 0
        if root_mtime > tracker_mtime:
            import shutil
            shutil.copy2(ROOT_TRACKER, TRACKER_FILE)

    if TRACKER_FILE.exists() and TRACKER_FILE.stat().st_size > 1024:
        tracker_mtime  = TRACKER_FILE.stat().st_mtime
        internal_mtime = MAIN_FILE.stat().st_mtime if MAIN_FILE.exists() else 0
        if tracker_mtime > internal_mtime:
            try:
                _parse_and_save_tracker(TRACKER_FILE)
            except Exception:
                if not MAIN_FILE.exists():
                    _create_sample_excel()
    elif not MAIN_FILE.exists():
        _create_sample_excel()


def _safe_date(v) -> str:
    if v is None or v is pd.NaT:
        return "TBD"
    s = str(v).strip()
    if not s or s.upper() in ("TBD", "NAN", "NAT", "NONE", "N/A"):
        return "TBD"
    try:
        ts = pd.to_datetime(v, errors="coerce")
        if pd.notna(ts):
            return ts.date().isoformat()
    except Exception:
        pass
    return s


def load_projects() -> pd.DataFrame:
    ensure_data_file()
    df = pd.read_excel(MAIN_FILE, sheet_name="Projects")
    for col in ("start_date", "target_date"):
        if col in df.columns:
            df[col] = df[col].apply(_safe_date)
    if "status" in df.columns:
        df["status"] = df["status"].fillna("In Progress").replace("", "In Progress")
    return df


def save_projects(df: pd.DataFrame) -> None:
    ensure_data_file()
    with pd.ExcelWriter(MAIN_FILE, engine="openpyxl", mode="a",
                        if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name="Projects", index=False)


def load_non_automatable(project_id: Optional[str] = None) -> pd.DataFrame:
    ensure_data_file()
    df = pd.read_excel(MAIN_FILE, sheet_name="Non_Automatable")
    if project_id:
        df = df[df["project_id"] == project_id].reset_index(drop=True)
    return df


def save_non_automatable(project_id: str, df: pd.DataFrame) -> None:
    ensure_data_file()
    all_df  = pd.read_excel(MAIN_FILE, sheet_name="Non_Automatable")
    other   = all_df[all_df["project_id"] != project_id]
    updated = pd.concat([other, df], ignore_index=True)
    with pd.ExcelWriter(MAIN_FILE, engine="openpyxl", mode="a",
                        if_sheet_exists="replace") as writer:
        updated.to_excel(writer, sheet_name="Non_Automatable", index=False)


def load_day_plan(project_id: Optional[str] = None) -> pd.DataFrame:
    ensure_data_file()
    df = pd.read_excel(MAIN_FILE, sheet_name="Day_Plan")
    df["date"] = pd.to_datetime(df["date"], errors="coerce").dt.date
    if project_id:
        df = df[df["project_id"] == project_id].reset_index(drop=True)
    return df


def save_day_plan(project_id: str, df: pd.DataFrame) -> None:
    ensure_data_file()
    all_df  = pd.read_excel(MAIN_FILE, sheet_name="Day_Plan")
    other   = all_df[all_df["project_id"] != project_id]
    updated = pd.concat([other, df], ignore_index=True)
    with pd.ExcelWriter(MAIN_FILE, engine="openpyxl", mode="a",
                        if_sheet_exists="replace") as writer:
        updated.to_excel(writer, sheet_name="Day_Plan", index=False)


def load_completion_plan() -> pd.DataFrame:
    ensure_data_file()
    return pd.read_excel(MAIN_FILE, sheet_name="Completion_Plan")


def save_completion_plan(df: pd.DataFrame) -> None:
    ensure_data_file()
    with pd.ExcelWriter(MAIN_FILE, engine="openpyxl", mode="a",
                        if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name="Completion_Plan", index=False)


def process_uploaded_file(uploaded_file):
    """
    Accept the Automation tracker.xlsx (multi-section sheets) OR
    the dashboard's own format (Projects / Non_Automatable / Day_Plan / Completion_Plan).
    """
    try:
        suffix = Path(uploaded_file.name).suffix.lower()

        if suffix == ".csv":
            df = pd.read_csv(uploaded_file)
            if "automated" in df.columns and "id" in df.columns:
                existing = load_projects()
                for _, row in df.iterrows():
                    pid = str(row.get("id", "")).strip()
                    if pid in existing["id"].values:
                        for col in ["automated","in_progress","non_automatable","status","notes"]:
                            if col in row and col in existing.columns:
                                existing.loc[existing["id"] == pid, col] = row[col]
                save_projects(existing)
                return True, f"Updated {len(df)} project rows from CSV."
            return False, "CSV must have 'id' and 'automated' columns."

        xl = pd.ExcelFile(uploaded_file)

        # ── Detect tracker format: any sheet with "Automation Status Summary" header ──
        tracker_sheets = [s for s in xl.sheet_names
                          if _has_tracker_format(xl.parse(s, header=None))]

        if tracker_sheets:
            if hasattr(uploaded_file, "getvalue"):
                raw = uploaded_file.getvalue()
            else:
                uploaded_file.seek(0)
                raw = uploaded_file.read()
            with open(TRACKER_FILE, "wb") as f:
                f.write(raw)
            _parse_and_save_tracker(TRACKER_FILE)
            return True, f"✅ Tracker imported ({', '.join(tracker_sheets)}). Dashboard updated."

        # ── Dashboard internal format ─────────────────────────────────────────
        internal_sheets = {"Projects", "Non_Automatable", "Day_Plan", "Completion_Plan"}
        msgs = []
        for sheet in xl.sheet_names:
            if sheet in internal_sheets:
                df = xl.parse(sheet)
                with pd.ExcelWriter(MAIN_FILE, engine="openpyxl", mode="a",
                                    if_sheet_exists="replace") as writer:
                    df.to_excel(writer, sheet_name=sheet, index=False)
                msgs.append(f"{sheet}: {len(df)} rows")
        if msgs:
            return True, "Imported: " + " | ".join(msgs)
        return False, (
            f"No matching sheets found. Sheets in file: {xl.sheet_names}. "
            f"Expected sheets with 'Automation Status Summary' header or internal format."
        )

    except Exception as e:
        return False, f"Import failed: {e}"


def get_tracker_download() -> bytes:
    """Return tracker xlsx for download (original if available)."""
    if TRACKER_FILE.exists() and TRACKER_FILE.stat().st_size > 1024:
        return TRACKER_FILE.read_bytes()

    # Reconstruct from internal data
    ensure_data_file()
    proj = pd.read_excel(MAIN_FILE, sheet_name="Projects")
    non  = pd.read_excel(MAIN_FILE, sheet_name="Non_Automatable")
    plan = pd.read_excel(MAIN_FILE, sheet_name="Day_Plan")
    comp = pd.read_excel(MAIN_FILE, sheet_name="Completion_Plan")

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        fw_pids = ["1p", "3p_wc", "3p_ltct"]
        _write_tracker_sheet(writer, "Firmware",
            proj[proj["id"].isin(fw_pids)],
            non[non["project_id"].isin(fw_pids)],
            plan[plan["project_id"].isin(fw_pids)],
            comp[comp["project_id"].isin(fw_pids + ["firmware_overall"])])

        for pid in proj["id"]:
            if pid in fw_pids:
                continue
            _write_tracker_sheet(writer, pid,
                proj[proj["id"] == pid],
                non[non["project_id"] == pid],
                plan[plan["project_id"] == pid],
                comp[comp["project_id"] == pid])
    return buf.getvalue()


def _write_tracker_sheet(writer, sheet_name, proj_df, non_df, plan_df, comp_df):
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils  import get_column_letter

    wb = writer.book
    ws = wb.create_sheet(sheet_name)

    hdr_fill  = PatternFill("solid", fgColor="0A3690")
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    sec_fill  = PatternFill("solid", fgColor="1645A4")
    sec_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    norm_font = Font(name="Calibri", size=10)

    row = 1

    def _write_section(title):
        nonlocal row
        ws.cell(row, 1, title).font = sec_font
        ws.cell(row, 1).fill = sec_fill
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

    def _write_df(df, cols):
        nonlocal row
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row, ci, col)
            c.font = hdr_font; c.fill = hdr_fill
        row += 1
        for _, r in df.iterrows():
            for ci, col in enumerate(cols, 1):
                val = r.get(col, "")
                if val is None or (isinstance(val, float) and val != val):
                    val = ""
                ws.cell(row, ci, val).font = norm_font
            row += 1
        row += 1

    _write_section("Automation Status Summary")
    avail = [c for c in ["id","name","total_cases","automatable","non_automatable","automated","in_progress","status"] if c in proj_df.columns]
    _write_df(proj_df[avail], avail)

    _write_section("Non-Automatable Test Cases")
    na_cols = [c for c in ["module","count","reason","approach"] if c in non_df.columns]
    _write_df(non_df[na_cols] if na_cols else non_df, na_cols or list(non_df.columns))

    _write_section("Day-by-Day Automation Plan")
    dp_cols = [c for c in ["date","project_id","module","actual_cases","cumulative","assigned_to","remarks","status"] if c in plan_df.columns]
    _write_df(plan_df[dp_cols] if dp_cols else plan_df, dp_cols or list(plan_df.columns))

    _write_section("Project Completion Plan")
    cp_cols = [c for c in ["project_id","name","total_cases","automatable","duration_days","daily_avg","start_date","expected_completion","status"] if c in comp_df.columns]
    _write_df(comp_df[cp_cols] if cp_cols else comp_df, cp_cols or list(comp_df.columns))

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 40)


def get_template_excel() -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pd.DataFrame(columns=["id","name","total_cases","automatable","non_automatable",
                               "automated","in_progress","team_size","start_date",
                               "target_date","daily_avg","status","priority","color","notes"]
            ).to_excel(writer, sheet_name="Projects", index=False)
        pd.DataFrame(columns=["project_id","module","count","reason","approach"]
            ).to_excel(writer, sheet_name="Non_Automatable", index=False)
        pd.DataFrame(columns=["project_id","date","module","planned_cases","actual_cases",
                               "cumulative","assigned_to","remarks","status"]
            ).to_excel(writer, sheet_name="Day_Plan", index=False)
        pd.DataFrame(columns=["project_id","name","total_cases","automatable","duration_days",
                               "daily_avg","start_date","expected_completion","status"]
            ).to_excel(writer, sheet_name="Completion_Plan", index=False)
    return buf.getvalue()
